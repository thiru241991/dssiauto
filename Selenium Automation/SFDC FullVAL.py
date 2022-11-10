from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.chrome.options import Options as ChromeOptions
import time
from time import sleep
import datetime
from selenium.common import exceptions
import openpyxl
from openpyxl.workbook import Workbook as workbook
from openpyxl.styles import Font, Alignment
import pyautogui
#https://bostonharborconsulting.atlassian.net/browse/AGIMONARCH-18050


excelpath = r"./SFDC FullVAL.xlsx"
workbook = openpyxl.load_workbook(excelpath)
sheet = workbook.create_sheet("sheet1", 0)
current_date_time = datetime.datetime.now()
date_time = current_date_time.strftime("%d-%m-%Y  %H hr %M min")
sheet.title = str(date_time)
sheet = workbook[str(date_time)]
sheet.column_dimensions['A'].width = 50
sheet.column_dimensions['B'].width = 15
sheet.cell(row=1,column=1).value = "Actions"
sheet.cell(row=1,column=2).value = "Time taken"

username = "jeganeedhi@bostonharborconsulting_prod.com.fullval"
password = "Jk12342022"
edge_options = EdgeOptions()
edge_options.add_argument("start-maximized")
driver = webdriver.Edge(options=edge_options, executable_path=r"./msedgedriver.exe")

driver.get("https://biotek--fullval.my.salesforce.com/home/home.jsp")
driver.find_element(By.ID,"username").send_keys(username)
driver.find_element(By.ID,"password").send_keys(password)
driver.find_element(By.ID,"Login").click()
def wait():
    time.sleep(3)
wait()
Oppty = driver.find_element(By.XPATH,'//*[@id="Opportunity_Tab"]').click()
#CPQQuotes = driver.find_element(By.XPATH,'//*[@id="01r1O000001lrV2_Tab"]/a').click()
#driver.get("https://biotek--fullval--c.sandbox.vf.force.com/apex/QTELST?sfdc.tabName=01r1O000001lrV2")
wait()
Opptyopen = driver.find_element(By.XPATH,'//a[contains(text(),"BHC Test Opp USA - DO NOT DELETE")]').click()
wait()
CreateQuote = driver.find_element(By.XPATH,'(//input[@name="biotek_create_quote"])[1]').click()
CPQiframe = driver.find_element(By.XPATH,'//iframe[@id="iframe1"]')
driver.switch_to.frame(CPQiframe)
print("switched to CPQ iframe")

wait()
def refreshwait():
    time.sleep(5)
    i = 0
    while i==0:
        try:   
            if (str(driver.find_element(By.XPATH,'//*[@id="head_highlight"]/div[2]/ul/li[1]').is_displayed()) == 'True'):
                i = 1
        except:
            exceptions.StaleElementReferenceException
            pass
refreshwait()
time.sleep(8)
iframeBody = driver.find_element(By.XPATH,'/html/body')
ActionChains(driver).move_to_element(iframeBody).context_click(iframeBody).perform()
wait()
pyautogui.press('up')
pyautogui.press('up')
pyautogui.press('enter')

def simpleproductwait():
    time.sleep(5)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[@id="progressbar_disc"]').get_attribute("style")) == 'display: none;' and str(driver.find_element(By.XPATH,'//*[@id="TH_ACTIONS"]/div[2]/button').is_displayed()) == 'True'):
            #print('simple product if')
            i = 1
simpleproductwait()
wait()
simpleproduct = driver.find_element(By.XPATH,'//*[@id="required_1863277"]').click()
wait()
simpleproductstart = time.time()
addsimpleproduct = driver.find_element(By.XPATH,'(//button[@id="BTN_add_quote"])[3]').click()

def addproductwait():
    time.sleep(10)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//button[@id="addproducts"]').is_displayed()) == 'True'):
            #print('add product if')
            i = 1

salesorgid = driver.find_element(By.XPATH, '//*[@id="Primary_Panel"]/table/tbody/tr[2]/td[9]').text
print(salesorgid)

if salesorgid == '':    
    driver.refresh()
    addproductwait()
    addproductstart = time.time()
    addproduct = driver.find_element(By.XPATH,'//button[@id="addproducts"]').click()

def bulkaddbtnwait():
    time.sleep(5)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[@id="progressbar_disc"]').get_attribute("style")) == 'display: none;' and str(driver.find_element(By.XPATH,'//button[@id="bulkaddprd"]').is_displayed()) == 'True'):
            #print('add product if')
            i = 1
bulkaddbtnwait()
simpleproductend = time.time()
addsimpleproducttime = simpleproductend - simpleproductstart
print('simple product time='+str(addsimpleproducttime))
sheet.cell(row=3,column=1).value = "Adding a Simple product"
sheet.cell(row=3,column=2).value = addsimpleproducttime
simpleproduct_SS = driver.save_screenshot("simpleproduct_SS.png")
addconfigproduct = driver.find_element(By.XPATH,'//*[@id="addproducts"]').click()

def prdsearchwait():
    time.sleep(5)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[@id="progressbar_disc"]').get_attribute("style")) == 'display: none;' and str(driver.find_element(By.XPATH,'//*[@id="TH_ACTIONS"]/div[2]').is_displayed()) == 'True'):
            #print('prdsearch if')
            i = 1
prdsearchwait()
prdsearch = driver.find_element(By.XPATH,'//input[@id="PART"]').click()
partno = "800TSI-SI"
wait()
configprdsearch = driver.find_element(By.XPATH,'//input[@id="PART"]').send_keys(partno)
wait()
partnosearch = driver.find_element(By.XPATH,'//*[@id="TH_ACTIONS"]/div[2]/button').click()

def config_prdwait():
    time.sleep(5)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//td[@id="TD_PartNumber"]//a[1]').is_displayed()) == 'True'):
            #print('config_prd if')
            i = 1
config_prdwait()

config_prdstart = time.time()
config_prd = driver.find_element(By.XPATH,'//td[@id="TD_PartNumber"]//a[1]').click()

def getpricingwait():
    time.sleep(5)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//button[@id="get_pricing"]').is_displayed()) == 'True'):
            #print('getpricing if')
            i = 1
getpricingwait()
config_prdend = time.time()
print('config_prd time='+str(config_prdend - config_prdstart))
wait()
getpricing = driver.find_element(By.XPATH,'//button[@id="get_pricing"]').click()

def saveproductwait():
    time.sleep(2)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[@id="cat_configure_save"]').is_displayed()) == 'True'):
            #print('saveproduct if')
            i = 1
saveproductwait()
wait()
saveproductstart = time.time()
saveproduct = driver.find_element(By.XPATH,'//*[@id="cat_configure_save"]').click()

def netvaluewait():
    time.sleep(5)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[@id="progressbar_disc"]').get_attribute("style")) == 'display: none;' and str(driver.find_element(By.XPATH,'//*[@id="seginnerbnr"]/div[7]/div[2]').is_displayed()) == 'True'):
            #print('netvalue if')
            i = 1
netvaluewait()
saveproductend = time.time()
saveconfigproducttime = saveproductend - saveproductstart
print('config_prd save time='+str(saveconfigproducttime))
sheet.cell(row=4,column=1).value = "Adding a Configurable product"
sheet.cell(row=4,column=2).value = saveconfigproducttime

netvalue = driver.find_element(By.XPATH, '//*[@id="seginnerbnr"]/div[7]/div[2]').text
print(netvalue)

bulkaddbtnwait()
bulkaddbtn = driver.find_element(By.XPATH,'//button[@id="bulkaddprd"]').click()

def bulkaddwait():
    time.sleep(2)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[@id="bulkaddsave"]').is_displayed()) == 'True'):
            #print('bulkadd if')
            i = 1
bulkaddwait()
bulkpartnos = driver.find_element(By.XPATH,'//textarea[@id="partnumbers"]').click()
partnos = "1020508,TAKE3,GEN5,8040501,1225101,7170011,1450541,1213032, 7082249,48294,8040501"
wait()
insertpartnos = driver.find_element(By.XPATH,'//textarea[@id="partnumbers"]').send_keys(partnos)
wait()
bulkaddsavestart = time.time()
bulkaddsave = driver.find_element(By.XPATH,'//*[@id="bulkaddsave"]').click()

bulkaddbtnwait()
bulkaddsaveend = time.time()
bulkadd10products = bulkaddsaveend - bulkaddsavestart
print('bulk add time='+str(bulkadd10products))
sheet.cell(row=5,column=1).value = "Bulk Add products"
sheet.cell(row=5,column=2).value = bulkadd10products
wait()
discdropdown =  Select(driver.find_element(By.XPATH,'//div[@class="row prd_container"]//select[1]'))
wait()
discdropdown.select_by_visible_text('MON Discount View')
#mondiscview = driver.find_element(By.XPATH,'/html/body/div[9]/div[3]/div/div/div/div[4]/div[7]/select/option[4]').click()

def YA9Idiscwait():
    time.sleep(2)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//tr[1]/td/a[contains(@id,"YA9I")]').is_displayed()) == 'True'):
            #print('bulkadd if')
            i = 1
YA9Idiscwait()
mondiscallcheck = driver.find_element(By.XPATH,'//input[@id="theadDiscountFirstChkbox"]').click()
wait()
YA9Idisc = driver.find_element(By.XPATH,'//tr[1]/td/a[contains(@id,"YA9I")]').click()
wait()
selectalldisc = driver.find_element(By.XPATH,'//span[text()="The record clicked"]/following::input').click()
wait()
discpercent = driver.find_element(By.XPATH,'(//span[text()="YA9I"])/following::div[2]/input')
discpercent.click()
wait()
#discpercentageclear = driver.find_element(By.XPATH,'/html/body/div[9]/div[21]/div/div/div[2]/div[1]/div[2]/div/input').clear()
discpercent.clear()
percentage = "10.0"
wait()
#discpercentage = driver.find_element(By.XPATH,'/html/body/div[9]/div[21]/div/div/div[2]/div[1]/div[2]/div/input').send_keys(percentage)
discpercent.send_keys(percentage)
wait()
discpercentsavestart = time.time()
discpercentsave = driver.find_element(By.XPATH,'(//button[contains(@class,"btn btn-list-cust")]/following-sibling::button)[3]').click()

bulkaddbtnwait()
discpercentsaveend = time.time()
Applydiscount = discpercentsaveend - discpercentsavestart
print('applying discount time='+str(Applydiscount))
sheet.cell(row=6,column=1).value = "Applying Bulk Discount"
sheet.cell(row=6,column=2).value = Applydiscount
wait()
netvalue = driver.find_element(By.XPATH, '//abbr[@id="tot_quote_netValue"]').text
print(netvalue)

def approvalnodewait():
    time.sleep(10)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[@id="progressbar_disc"]').get_attribute("style")) == 'display: none;' and str(driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/approval_node.svg")]').is_displayed()) == 'True'):
            #print('approval node if')
            i = 1

if netvalue == '0.00':    
    reprice = driver.find_element(By.XPATH,'//*[@id="btn_Reprice"]').click()
    approvalnodewait()

approvalnodewait()
        
approvalnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/approval_node.svg")]').click()

def approvaliframewait():
    time.sleep(3)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[@id="approvalIframe"]').is_displayed()) == 'True'):
            #print('approval iframe if')
            i = 1
approvaliframewait()

approvaliframe = driver.find_element(By.XPATH, '//*[@id="approvalIframe"]')
driver.switch_to.frame(approvaliframe)
print("switched to approval iframe")

def justifywait():
    time.sleep(10)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'(//textarea[@id="JUSTIFICATION_COMMENTS"])[2]').is_displayed()) == 'True'):
            #print('justify if')
            i = 1
justifywait()
j_comment = driver.find_element(By.XPATH,'(//textarea[@id="JUSTIFICATION_COMMENTS"])[2]').click()
comment = "automation test"
wait()
justify = driver.find_element(By.XPATH,'(//textarea[@id="JUSTIFICATION_COMMENTS"])[2]').send_keys(comment)

def approvebuttonwait():
    time.sleep(2)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;')  and str(driver.find_element(By.XPATH,'//*[@id="approval_act_btn"]').is_displayed()) == 'True'):
            #print('approve button if')
            i = 1
approvebuttonwait()
approvebuttonstart = time.time()
approvebutton = driver.find_element(By.XPATH,'//*[@id="approval_act_btn"]')
ActionChains(driver).move_to_element(approvebutton).click(approvebutton).perform()

def approvecompletewait():
    time.sleep(5)
    i = 0
    while i==0:
        try:
            if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'/html/body/div[9]/div[2]/div[2]/table/tbody/tr/td/div').is_displayed()) == 'True'):
                i = 1
        except:
            exceptions.StaleElementReferenceException
            pass 
approvecompletewait()
approvebuttonend = time.time()
ApprovingQuote = approvebuttonend - approvebuttonstart
print('approve time='+str(ApprovingQuote))
sheet.cell(row=7,column=1).value = "Self Approving Quote"
sheet.cell(row=7,column=2).value = ApprovingQuote
approvecompletewait()
driver.switch_to.default_content()
print('switched to default content')

def ordersnodewait():
    time.sleep(10)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/order.svg")]').is_displayed()) == 'True'):
            print('Ordernode if')
            i = 1
ordersnodewait()
WebDriverWait(driver,60).until(EC.presence_of_element_located((By.XPATH, '//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/order.svg")]')))
ordersnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/order.svg")]')
ActionChains(driver).move_to_element(ordersnode).click(ordersnode).perform()
print('clicked orders node')

def createRFOwait():
    time.sleep(5)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//button[text()="CREATE RFO"]').is_displayed()) == 'True'):
            #print('Create RFO if')
            i = 1
createRFOwait()

createRFO = driver.find_element(By.XPATH,'//button[text()="CREATE RFO"]').click()

def POnumberwait():
    time.sleep(5)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//input[@id="po_number"]').is_displayed()) == 'True'):
            #print('PO number if')
            i = 1
ponum = "0987"
POnumberwait()
PO_number = driver.find_element(By.XPATH,'//input[@id="po_number"]').send_keys(ponum)
POnumberwait()
datepicker = driver.find_element(By.XPATH,'//*[@id="po_date"]').click()
wait()
datepicker = driver.find_element(By.XPATH,'/html/body/div[11]/div[1]/table/tbody/tr[3]/td[2]').click()
#print('date picked')
wait()
missingreason = 'automation test'
PO_missing = driver.find_element(By.XPATH,'//input[@id="po_file_missing"]').send_keys(missingreason)
POnumberwait()
CreateRFOsubmitstart = time.time()
CreateRFOsubmit = driver.find_element(By.XPATH,'//button[@id="CreateRFO"]').click()

def BacktoQuotewait():
    time.sleep(10)
    i = 0
    while i==0:
        try:
            if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//p/a[@class="quote-title-span7"]').is_displayed()) == 'True'):
                #print('BacktoQuotes if')
                i = 1
        except:
            exceptions.StaleElementReferenceException
            pass  
BacktoQuotewait()
CreateRFOsubmitend = time.time()
CreatingRFO = CreateRFOsubmitend - CreateRFOsubmitstart
print('Create RFO time='+str(CreatingRFO))
sheet.cell(row=8,column=1).value = "Creating RFO"
sheet.cell(row=8,column=2).value = CreatingRFO
wait()
BacktoQuote = driver.find_element(By.XPATH,'//*[contains(@onclick, "gotquote_num_rev(this.innerHTML);")]').click()

def revisionsnodewait():
    time.sleep(10)
    i = 0
    while i==0:
        try:
            if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/Revisions_node.svg")]').is_displayed()) == 'True'):
                #print('revisionsnode if')
                i = 1
        except:
            exceptions.StaleElementReferenceException
            pass
revisionsnodewait()

revisionsnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/Revisions_node.svg")]').click()

def NewRevisionwait():
    time.sleep(10)
    i = 0
    while i==0:
            if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//button[@id="Newrevision"]').is_displayed()) == 'True'):
                #print('NewRevision if')
                i = 1
NewRevisionwait()
NewRevisionstart = time.time()
NewRevision = driver.find_element(By.XPATH,'//button[@id="Newrevision"]').click()
NewRevisionwait()
NewRevisionend = time.time()
Creating_New_Revision = NewRevisionend - NewRevisionstart
print('New Revision time='+str(Creating_New_Revision))
sheet.cell(row=9,column=1).value = "Creating New Revision"
sheet.cell(row=9,column=2).value = Creating_New_Revision

def productsnodewait():
    time.sleep(5)
    i = 0
    while i==0:
        if ((str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(driver.find_element(By.XPATH,'//*[@id="progressbar_disc"]').get_attribute("style")) == 'display: none;' and str(driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/product.svg")]').is_displayed()) == 'True'):
            #print('productsnode if')
            i = 1
productsnodewait()
productsnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/product.svg")]').click()

def deletechkboxwait():
    time.sleep(5)
    i = 0
    while i==0:
        if (str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;' and str(driver.find_element(By.XPATH,'//button[@id="btn_bulkDelete"]').is_displayed()) == 'True'):
            #print('deletechkbox if')
            i = 1
deletechkboxwait()
#wait()
#driver.execute_script("window.scrollTo(0, 10)")
wait()
prdpagination = driver.find_element(By.XPATH,'//div[@id="itemsPerPageDropdown"]').click()
wait()
prdpagination = driver.find_element(By.XPATH,'//a[contains(text(),"20")]').click()
deletechkboxwait()
deletechkbox = driver.find_element(By.XPATH,'//input[@class="custom inp_chkHdrMulDel"]').click()

time.sleep(5)
deletebutton = driver.find_element(By.XPATH,'//button[@id="btn_bulkDelete"]').click()

time.sleep(5)
deleteconfirmstart = time.time()
deleteconfirm = driver.find_element(By.XPATH,'//button[text()="DELETE"]').click()
#print('deleted')
addproductwait()
deleteconfirmend = time.time()
Bulkdelete12products = deleteconfirmend - deleteconfirmstart
print('bulk delete time='+str(Bulkdelete12products))
sheet.cell(row=10,column=1).value = "Bulk Deleting Products"
sheet.cell(row=10,column=2).value = Bulkdelete12products

sheet['A1'].font = Font(bold=True, size = 12)
sheet['A1'].alignment = Alignment(horizontal='center')
sheet['B1'].font = Font(bold=True, size = 12)
sheet['B1'].alignment = Alignment(horizontal='center')
'''for row in sheet['B']:
    cell = row[:10]
    cell.alignment = Alignment(horizontal='center')'''

workbook.save(excelpath)
time.sleep(10)
driver.quit()
