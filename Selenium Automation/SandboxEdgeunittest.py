import SandboxEdgeunittest
from HtmlTestRunner.runner import HTMLTestRunner
import unittest
from unittest.main import MODULE_EXAMPLES
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.common import exceptions
import time
import datetime
from time import sleep
import openpyxl
from openpyxl.workbook import Workbook as workbook
from openpyxl.styles import Font, Alignment


class Automation(unittest.TestCase):
    def wait(self):
            time.sleep(3)
    def pageloader(self): 
        WebDriverWait(self.driver, 3600).until(EC.invisibility_of_element_located((By.XPATH, '//*[contains(@id,"progressbar_disc")]')))
        self.wait()
        WebDriverWait(self.driver, 3600).until(EC.invisibility_of_element_located((By.XPATH, '//*[contains(@id,"pageloader")]')))
        #print('pageloader func executed')
    def bulkaddbtnwait(self):
            self.pageloader()
            WebDriverWait(self.driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//button[@id="bulkaddprd"]')))
    def addproductwait(self):
            self.pageloader()
            WebDriverWait(self.driver,60000).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="addproducts"]')))

    @classmethod
    def setUpClass(inst):
        
        inst.excelpath = r".\\Sandbox.xlsx"
        inst.workbook = openpyxl.load_workbook(inst.excelpath)
        inst.sheet = inst.workbook.create_sheet("sheet1", 0)
        current_date_time = datetime.datetime.now()
        date_time = current_date_time.strftime("%d-%m-%Y  %H hr %M min")
        inst.sheet.title = str(date_time)
        inst.sheet = inst.workbook[str(date_time)]
        inst.sheet.column_dimensions['A'].width = 50
        inst.sheet.column_dimensions['B'].width = 15
        inst.sheet.cell(row=1,column=1).value = "Actions"
        inst.sheet.cell(row=1,column=2).value = "Time taken"

        inst.username = "saitejareddy"
        inst.password = "Sai@8994"
        inst.domain = "agilenttechnologiesinc_tst1"
        edge_options = EdgeOptions()
        edge_options.add_argument("start-maximized")
        inst.driver = webdriver.Edge(options=edge_options, executable_path=r".\\msedgedriver.exe")
        inst.driver.get("https://sandbox.webcomcpq.com/login.aspx")
        

    def test_001_login(self):
        self.driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_txtUsername").send_keys(self.username)
        self.driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_txtPassword").send_keys(self.password)
        self.driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_txtDomainName").send_keys(self.domain)
        self.driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_btnLogin").click()

        WebDriverWait(self.driver,3600).until(EC.presence_of_element_located((By.XPATH, '//*[@id="filllevel_container"]')))

        self.pageloader()
        allquotes = self.driver.find_element(By.XPATH,'//*[@id="fillallquote1"]/a')
        #print("Element is visible? " + str(allquotes.is_displayed()))
        ActionChains(self.driver).move_to_element(allquotes).click(allquotes).perform()
        self.pageloader()
        allquotesearch = self.driver.execute_script('return document.querySelector("#qterevall > div.search_button input")')
        self.driver.execute_script("arguments[0].click();", allquotesearch)
        self.driver.execute_script('return document.querySelector("#qterevall > div.search_button input").value = "05984507" ')
        #print('entered')
        
        allsearch = self.driver.find_element(By.XPATH,'//*[@id="Quote_tableall"]/thead/tr/th[1]/div[2]/button').click()

        
        
    def test_002_AddProduct(self):
        self.pageloader()
        self.wait()
        WebDriverWait(self.driver,3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="fillallquote1"]/a')))
        quoteclick = self.driver.find_element(By.XPATH,'/html/body/div[5]/div[1]/div[6]/div[2]/div[1]/div/div/div[2]/div/div[1]/table/tbody/tr/td[2]/a')
        self.driver.execute_script("arguments[0].click();", quoteclick)

        def refreshwait():
            time.sleep(5)
            i = 0
            while i==0:
                try:   
                    if (str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;'):
                        i = 1
                except:
                    exceptions.StaleElementReferenceException
                    pass
        refreshwait()
        self.driver.refresh()

        self.addproductwait()
        addproductstart = time.time()

        addproduct = self.driver.find_element(By.XPATH,'//*[@id="addproducts"]')
        self.driver.execute_script('arguments[0].click()', addproduct)
        self.pageloader()

        salesorgid = self.driver.find_element(By.XPATH, '//*[@id="QuoteTypePHP"]').text
        print(salesorgid)

        if salesorgid == '':    
            self.driver.refresh()
            self.addproductwait()
            addproductstart = time.time()
            #---!addproduct = driver.find_element(By.XPATH,'/html/body/div[9]/div[3]/div/div/div/div[4]/div[5]/button[1]').click()
            addproduct = self.driver.find_element(By.XPATH,'//*[@id="addproducts"]')
            self.driver.execute_script('arguments[0].click()', addproduct)
            
        def simpleproductwait():
            self.pageloader()
            WebDriverWait(self.driver,60000).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="TH_ACTIONS"]/div[2]/button')))
            
        simpleproductwait()
        addproductend = time.time()
        addproducttime = addproductend - addproductstart
        print('add product time='+str(addproducttime))
        self.sheet.cell(row=2,column=1).value = "Navigating to Product Catalog by Add Product"
        self.sheet.cell(row=2,column=2).value = addproducttime
        self.wait()

    
    def test_003_AddingSimpleproduct(self):
        driver=self.driver
        simpleproduct = driver.find_element(By.XPATH,'(//label[text()="Configurable"]/following::input)[2]')
        driver.execute_script("arguments[0].click();", simpleproduct)
        self.wait()
        time.sleep(3)
        simpleproductstart = time.time()
        addsimpleproduct = driver.find_element(By.XPATH,'(//button[@id="BTN_add_quote"])[3]').click()
        #print('add to cart')
        
        self.bulkaddbtnwait()

        self.pageloader()
        simpleproductend = time.time()
        addsimpleproducttime = simpleproductend - simpleproductstart
        print('simple product time='+str(addsimpleproducttime))
        self.sheet.cell(row=3,column=1).value = "Adding a Simple product"
        self.sheet.cell(row=3,column=2).value = addsimpleproducttime

    def test_004_ConfigurePrdSave(self):
        driver=self.driver
        addconfigproduct = driver.find_element(By.XPATH,'//*[@id="addproducts"]').click()

        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//input[@id="PART"]')))
        #addproduct.click()
        prdsearch = driver.find_element(By.XPATH,'//input[@id="PART"]')
        driver.execute_script("arguments[0].click();", prdsearch)
        partno = "800TSI-SI"
        self.wait()
        configprdsearch = driver.find_element(By.XPATH,'//input[@id="PART"]').send_keys(partno)
        self.wait()
        partnosearch = driver.find_element(By.XPATH,'//*[@id="TH_ACTIONS"]/div[2]/button').click()
        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//td[@id="TD_PartNumber"]//a[1]')))


        config_prdstart = time.time()
        config_prd = driver.find_element(By.XPATH,'//td[@id="TD_PartNumber"]//a[1]').click()

        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//button[@id="get_pricing"]')))
        config_prdend = time.time()
        print('config_prd time='+str(config_prdend - config_prdstart))
        self.wait()
        getpricing = driver.find_element(By.XPATH,'//button[@id="get_pricing"]').click()


        self.wait()
        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="cat_configure_save"]')))
        saveproductstart = time.time()
        saveproduct = driver.find_element(By.XPATH,'//*[@id="cat_configure_save"]').click()

        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="seginnerbnr"]/div[7]/div[2]')))
        saveproductend = time.time()
        saveconfigproducttime = saveproductend - saveproductstart
        print('config_prd save time='+str(saveconfigproducttime))
        self.sheet.cell(row=4,column=1).value = "Adding a Configurable product"
        self.sheet.cell(row=4,column=2).value = saveconfigproducttime

        netvalue = driver.find_element(By.XPATH, '//*[@id="seginnerbnr"]/div[7]/div[2]').text
        print(netvalue)

    def test_005_BulkAddPrd(self):
        driver=self.driver
        self.bulkaddbtnwait()
        bulkaddbtn = driver.find_element(By.XPATH,'//button[@id="bulkaddprd"]').click()

        #pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="bulkaddsave"]')))
        bulkpartnos = driver.find_element(By.XPATH,'//textarea[@id="partnumbers"]').click()
        partnos = "1020508,TAKE3,GEN5,8040501,1225101,7170011,1450541,1213032, 7082249,48294,8040501"
        self.wait()
        insertpartnos = driver.find_element(By.XPATH,'//textarea[@id="partnumbers"]').send_keys(partnos)
        self.wait()
        bulkaddsavestart = time.time()
        bulkaddsave = driver.find_element(By.XPATH,'//*[@id="bulkaddsave"]').click()

        self.bulkaddbtnwait()
        bulkaddsaveend = time.time()
        bulkadd10products = bulkaddsaveend - bulkaddsavestart
        print('bulk add time='+str(bulkadd10products))
        self.sheet.cell(row=5,column=1).value = "Bulk Add products"
        self.sheet.cell(row=5,column=2).value = bulkadd10products
        self.wait()

    def test_006_ApplyingBulkDiscount(self):
        driver=self.driver
        discdropdown =  Select(driver.find_element(By.XPATH,'//div[@class="row prd_container"]//select[1]'))
        self.wait()
        discdropdown.select_by_visible_text('MON Discount View')
        #mondiscview = driver.find_element(By.XPATH,'/html/body/div[9]/div[3]/div/div/div/div[4]/div[7]/select/option[4]').click()
        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//tr[1]/td/a[contains(@id,"YA9I")]')))
        mondiscallcheck = driver.find_element(By.XPATH,'//input[@id="theadDiscountFirstChkbox"]').click()
        self.wait()
        YA9Idisc = driver.find_element(By.XPATH,'//tr[1]/td/a[contains(@id,"YA9I")]').click()
        self.wait()
        selectalldisc = driver.find_element(By.XPATH,'//span[text()="The record clicked"]/following::input').click()
        self.wait()
        discpercent = driver.find_element(By.XPATH,'(//span[text()="YA9I"])/following::div[2]/input')
        discpercent.click()
        self.wait()
        #discpercentageclear = driver.find_element(By.XPATH,'/html/body/div[9]/div[21]/div/div/div[2]/div[1]/div[2]/div/input').clear()
        discpercent.clear()
        percentage = "10.0"
        self.wait()
        #discpercentage = driver.find_element(By.XPATH,'/html/body/div[9]/div[21]/div/div/div[2]/div[1]/div[2]/div/input')
        discpercent.send_keys(percentage)
        self.wait()
        discpercentsavestart = time.time()
        discpercentsave = driver.find_element(By.XPATH,'(//button[contains(@class,"btn btn-list-cust")]/following-sibling::button)[3]').click()
        self.bulkaddbtnwait()
        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//abbr[@id="tot_quote_netValue"]')))
        discpercentsaveend = time.time()
        Applydiscount = discpercentsaveend - discpercentsavestart
        print('applying discount time='+str(Applydiscount))
        self.sheet.cell(row=6,column=1).value = "Applying Bulk Discount"
        self.sheet.cell(row=6,column=2).value = Applydiscount
        self.wait()

    def test_007_SelfApprovingQuote(self):
        driver=self.driver
        netvalue = driver.find_element(By.XPATH, '//abbr[@id="tot_quote_netValue"]').text
        print(netvalue)

        def approvalnodewait():
            self.pageloader()
            WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/approval_node.svg")]')))

        if netvalue == '0.00':    
            reprice = driver.find_element(By.XPATH,'//*[@id="btn_Reprice"]').click()
            approvalnodewait()

        approvalnodewait()
                
        approvalnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/approval_node.svg")]').click()

        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="approvalIframe"]')))

        approvaliframe = driver.find_element(By.XPATH, '//*[@id="approvalIframe"]')
        self.wait()
        driver.switch_to.frame(approvaliframe)
        print("switched to approval iframe")


        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '(//textarea[@id="JUSTIFICATION_COMMENTS"])[2]')))
        j_comment = driver.find_element(By.XPATH,'(//textarea[@id="JUSTIFICATION_COMMENTS"])[2]').click()
        comment = "automation test"
        self.wait()
        justify = driver.find_element(By.XPATH,'(//textarea[@id="JUSTIFICATION_COMMENTS"])[2]').send_keys(comment)

        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//div[2]/button[1][@id="approval_act_btn"]')))
        approvebuttonstart = time.time()
        approvebutton = driver.find_element(By.XPATH,'//div[2]/button[1][@id="approval_act_btn"]')
        ActionChains(driver).move_to_element(approvebutton).click(approvebutton).perform()


        self.pageloader()
        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="approvalIframe"]')))
        approvebuttonend = time.time()
        ApprovingQuote = approvebuttonend - approvebuttonstart
        print('approve time='+str(ApprovingQuote))
        self.sheet.cell(row=7,column=1).value = "Self Approving Quote"
        self.sheet.cell(row=7,column=2).value = ApprovingQuote
        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="approvalIframe"]')))
        self.wait()
        driver.switch_to.default_content()
        self.wait()
        print('switched to default content')
        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//img[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/order.svg")]')))

    def test_008_CreatingRFO(self):
        driver=self.driver
        ordersnode = driver.find_element(By.XPATH,'//img[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/order.svg")]')
        #ActionChains(driver).move_to_element(ordersnode).click().perform()
        driver.execute_script("arguments[0].click();", ordersnode)
        print('clicked orders node')

        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//button[text()="CREATE RFO"]')))
        createRFO = driver.find_element(By.XPATH,'//button[text()="CREATE RFO"]').click()

        def POnumberwait():
            self.pageloader()
            WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="po_number"]')))

        POnumberwait()
        ponum = "0987"
        PO_number = driver.find_element(By.XPATH,'//input[@id="po_number"]').send_keys(ponum)
        POnumberwait()
        datepicker = driver.find_element(By.XPATH,'//input[@id="po_date"]').click()
        self.wait()
        datepicker = driver.find_element(By.XPATH,'//tr[3]/td[text()="10"]').click()
        #print('date picked')
        self.wait()
        missingreason = "automation test"
        PO_missing = driver.find_element(By.XPATH,'//input[@id="po_file_missing"]').send_keys(missingreason)
        POnumberwait()
        CreateRFOsubmitstart = time.time()
        CreateRFOsubmit = driver.find_element(By.XPATH,'//button[@id="CreateRFO"]').click()

        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//p/a[@class="quote-title-span7"]')))
        CreateRFOsubmitend = time.time()
        CreatingRFO = CreateRFOsubmitend - CreateRFOsubmitstart
        print('Create RFO time='+str(CreatingRFO))
        self.sheet.cell(row=8,column=1).value = "Creating RFO"
        self.sheet.cell(row=8,column=2).value = CreatingRFO
        self.wait()
        BacktoQuote = driver.find_element(By.XPATH,'//*[@id="Primary_Panel"]/table/tbody/tr[2]/td[7]/p/a')
        driver.execute_script("arguments[0].click();", BacktoQuote)

        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/Revisions_node.svg")]')))
        self.wait()

    def test_009_CreatingNewRevision(self):
        driver=self.driver
        revisionsnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/Revisions_node.svg")]')
        driver.execute_script('arguments[0].scrollIntoView({block: "center"});', revisionsnode)
        driver.execute_script("arguments[0].click();", revisionsnode)

        def NewRevisionwait():
            self.pageloader()
            WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//button[@id="Newrevision"]')))
        NewRevisionwait()
        NewRevisionstart = time.time()
        NewRevision = driver.find_element(By.XPATH,'//button[@id="Newrevision"]').click()
        self.pageloader()
        WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/product.svg")]')))
        NewRevisionend = time.time()
        Creating_New_Revision = NewRevisionend - NewRevisionstart
        print('New Revision time='+str(Creating_New_Revision))
        self.sheet.cell(row=9,column=1).value = "Creating New Revision"
        self.sheet.cell(row=9,column=2).value = Creating_New_Revision
        self.wait()

    def test_010_BulkDeletingProducts(self):
        driver=self.driver
        productsnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/product.svg")]').click()

        #wait()
        #driver.execute_script("window.scrollTo(0, 10)")
        def deletechkboxwait():
            self.pageloader()
            WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//button[@id="btn_bulkDelete"]')))
        deletechkboxwait()
        self.wait()
        prdpagination = driver.find_element(By.XPATH,'//div[@id="itemsPerPageDropdown"]').click()
        self.wait()
        prdpagination = driver.find_element(By.XPATH,'//a[contains(text(),"20")]').click()
        deletechkboxwait()
        deletechkbox = driver.find_element(By.XPATH,'//input[@class="custom inp_chkHdrMulDel"]').click()

        time.sleep(5)
        deletebutton = driver.find_element(By.XPATH,'//button[@id="btn_bulkDelete"]').click()

        time.sleep(5)
        deleteconfirmstart = time.time()
        deleteconfirm = driver.find_element(By.XPATH,'//button[text()="DELETE"]').click()
        #print('deleted')
        self.addproductwait()
        deleteconfirmend = time.time()
        Bulkdelete12products = deleteconfirmend - deleteconfirmstart
        print('bulk delete time='+str(Bulkdelete12products))
        self.sheet.cell(row=10,column=1).value = "Bulk Deleting Products"
        self.sheet.cell(row=10,column=2).value = Bulkdelete12products

    def test_011_UpdatingExcel(self):
        driver=self.driver
        
        self.sheet['A1'].font = Font(bold=True, size = 12)
        self.sheet['A1'].alignment = Alignment(horizontal='center')
        self.sheet['B1'].font = Font(bold=True, size = 12)
        self.sheet['B1'].alignment = Alignment(horizontal='center')
        
        self.workbook.save(self.excelpath)
        time.sleep(10)
        self.driver.quit()

'''if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='Test Report2',report_title = 'Test Results',report_name = 'Test Report',open_in_browser=True))'''
 #from tests.test_1 import TestCase1
#from tests.test_2 import TestCase2


if __name__ == '__main__':
   
    test1 = unittest.TestLoader().loadTestsFromTestCase(SandboxEdgeunittest.Automation)
    #test2 = unittest.TestLoader().loadTestsFromTestCase(TestCase2)
    suite = unittest.TestSuite([test1])
    runner = HTMLTestRunner(output='Test Reports',report_title = 'Automation Test Results',report_name = 'Test Report',open_in_browser=True)
    runner.run(suite)