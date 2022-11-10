from calendar import c
import unittest
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.common import exceptions
import time
import datetime
from time import sleep
import openpyxl
from openpyxl.workbook import Workbook as workbook
from openpyxl.styles import Font, Alignment
import HtmlTestRunner

class Automation(unittest.TestCase):
    def wait(self):
            time.sleep(3)
    def pageloader(self): 
        WebDriverWait(self.driver, 3600).until(EC.invisibility_of_element_located((By.XPATH, '//*[contains(@id,"progressbar_disc")]')))
        self.wait()
        WebDriverWait(self.driver, 3600).until(EC.invisibility_of_element_located((By.XPATH, '//*[contains(@id,"pageloader")]')))
        print('pageloader func executed')
    @classmethod
    def setUpClass(inst):
        
        inst.excelpath = r".\\Sandbox.xlsx"
        inst.workbook = openpyxl.load_workbook(inst.excelpath)
        inst.sheet = inst.workbook.create_sheet("sheet1", 0)
        current_date_time = datetime.datetime.now()
        date_time = current_date_time.strftime("%d-%m-%Y  %H hr %M min")
        inst.sheet.title = str(date_time)
        inst.sheet = inst.workbook[str(date_time)]
        inst.sheet.column_dimensions['A'].width = 50
        inst.sheet.column_dimensions['B'].width = 15
        inst.sheet.cell(row=1,column=1).value = "Actions"
        inst.sheet.cell(row=1,column=2).value = "Time taken"

        inst.username = "saitejareddy"
        inst.password = "Sai@8994"
        inst.domain = "agilenttechnologiesinc_tst1"
        edge_options = EdgeOptions()
        edge_options.add_argument("start-maximized")
        inst.driver = webdriver.Edge(options=edge_options, executable_path=r".\\msedgedriver.exe")
        inst.driver.get("https://sandbox.webcomcpq.com/login.aspx")

    def test_regression_170899(self):
        self.driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_txtUsername").send_keys(self.username)
        self.driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_txtPassword").send_keys(self.password)
        self.driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_txtDomainName").send_keys(self.domain)
        self.driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_btnLogin").click()

        WebDriverWait(self.driver,3600).until(EC.presence_of_element_located((By.XPATH, '//*[@id="filllevel_container"]')))

        def allqoutewait():
            i = 0
            while i==0:
                if ((str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(self.driver.find_element(By.XPATH,'//*[@id="fillallquote1"]/a').is_displayed()) == 'True'):
                    #print('quote if')
                    i = 1
        allqoutewait()
        allquotes = self.driver.find_element(By.XPATH,'//*[@id="fillallquote1"]/a')
        #print("Element is visible? " + str(allquotes.is_displayed()))
        ActionChains(self.driver).move_to_element(allquotes).click(allquotes).perform()
        def quotesearch():
            def SearchingQuote():
                time.sleep(5)
                i = 0
                while i==0:
                    if ((str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(self.driver.find_element(By.XPATH,'/html/body/div[5]/div[1]/div[6]/div[2]/div[1]/div/div/div[2]/div/div[1]/table/thead/tr/th[2]/div[2]/input').is_displayed()) == 'True'):
                        #print('quotesearch if')
                        i = 1
            SearchingQuote()
            allquotesearch = self.driver.execute_script('return document.querySelector("#qterevall > div.search_button input")')
            self.driver.execute_script("arguments[0].click();", allquotesearch)
            self.driver.execute_script('return document.querySelector("#qterevall > div.search_button input").value = "05984507" ')
            #print('entered')
        quotesearch()
        allsearch = self.driver.find_element(By.XPATH,'//*[@id="Quote_tableall"]/thead/tr/th[1]/div[2]/button').click()

        
        
    def test_regression_18909(self):
        self.pageloader()
        self.wait()
        WebDriverWait(self.driver,3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="fillallquote1"]/a')))
        quoteclick = self.driver.find_element(By.XPATH,'/html/body/div[5]/div[1]/div[6]/div[2]/div[1]/div/div/div[2]/div/div[1]/table/tbody/tr/td[2]/a')
        self.driver.execute_script("arguments[0].click();", quoteclick)

        def refreshwait():
            time.sleep(5)
            i = 0
            while i==0:
                try:   
                    if (str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;'):
                        i = 1
                except:
                    exceptions.StaleElementReferenceException
                    pass
        refreshwait()
        self.driver.refresh()

        def addproductwait():
            WebDriverWait(self.driver,60000).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="addproducts"]')))
        addproductwait()
        addproductstart = time.time()

        addproduct = self.driver.find_element(By.XPATH,'//*[@id="addproducts"]')
        self.driver.execute_script('arguments[0].click()', addproduct)
        def bulkaddbtnwait():
            time.sleep(5)
            i = 0
            while i==0:
                if ((str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(self.driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;') and str(self.driver.find_element(By.XPATH,'//*[@id="progressbar_disc"]').get_attribute("style")) == 'display: none;' and str(self.driver.find_element(By.XPATH,'/html/body/div[9]/div[3]/div/div/div/div[4]/div[5]/button[2]').is_displayed()) == 'True'):
                    #print('add product if')
                    i = 1 

        salesorgid = self.driver.find_element(By.XPATH, '//*[@id="QuoteTypePHP"]').text
        print(salesorgid)

        if salesorgid == '':    
            self.driver.refresh()
            addproductwait()
            addproductstart = time.time()
            #addproduct = driver.find_element(By.XPATH,'/html/body/div[9]/div[3]/div/div/div/div[4]/div[5]/button[1]').click()
            addproduct = self.driver.find_element(By.XPATH,'//*[@id="addproducts"]')
            self.driver.execute_script('arguments[0].click()', addproduct)
            
        def simpleproductwait():
            WebDriverWait(self.driver,60000).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="TH_ACTIONS"]/div[2]/button')))
            
        simpleproductwait()
        addproductend = time.time()
        addproducttime = addproductend - addproductstart
        print('add product time='+str(addproducttime))
        self.sheet.cell(row=2,column=1).value = "Navigating to Product Catalog by Add Product"
        self.sheet.cell(row=2,column=2).value = addproducttime
        self.wait()
        
        self.sheet['A1'].font = Font(bold=True, size = 12)
        self.sheet['A1'].alignment = Alignment(horizontal='center')
        self.sheet['B1'].font = Font(bold=True, size = 12)
        self.sheet['B1'].alignment = Alignment(horizontal='center')
        '''for row in sheet['B']:
            cell = row[:10]
            cell.alignment = Alignment(horizontal='center')'''

        self.workbook.save(self.excelpath)
        time.sleep(10)
        self.driver.quit()
if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='Test Report'))