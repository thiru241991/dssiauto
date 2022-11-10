from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.common import exceptions
import time
import datetime
from time import sleep
import openpyxl
from openpyxl.workbook import Workbook as workbook
from openpyxl.styles import Font, Alignment

excelpath = r"./Sandbox.xlsx"
workbook = openpyxl.load_workbook(excelpath)
sheet = workbook.create_sheet("sheet1", 0)
current_date_time = datetime.datetime.now()
date_time = current_date_time.strftime("%d-%m-%Y  %H hr %M min")
sheet.title = str(date_time)
sheet = workbook[str(date_time)]
sheet.column_dimensions['A'].width = 50
sheet.column_dimensions['B'].width = 15
sheet.cell(row=1,column=1).value = "Actions"
sheet.cell(row=1,column=2).value = "Time taken"

username = "saitejareddy"
password = "Sai@8994"
domain = "agilenttechnologiesinc_tst1"
edge_options = EdgeOptions()
edge_options.add_argument("start-maximized")
driver = webdriver.Edge(options=edge_options, executable_path=r"./msedgedriver.exe")
driver.get("https://sandbox.webcomcpq.com/login.aspx")
driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_txtUsername").send_keys(username)
driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_txtPassword").send_keys(password)
driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_txtDomainName").send_keys(domain)
driver.find_element(By.ID,"ctl00_MainContentPlaceHolder_btnLogin").click()

WebDriverWait(driver,3600).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="filllevel_container"]')))

def wait():
    time.sleep(3)

def pageloader(): 
    WebDriverWait(driver, 3600).until(EC.invisibility_of_element_located((By.XPATH, '//*[contains(@id,"progressbar_disc")]')))
    wait()
    WebDriverWait(driver, 3600).until(EC.invisibility_of_element_located((By.XPATH, '//*[contains(@id,"pageloader")]')))
    
    '''if (driver.find_element(By.XPATH,'//*[contains(@id,"progressbar_disc")]')).is_displayed:
        WebDriverWait(driver, 3600).until(EC.invisibility_of_element_located((By.XPATH, '//*[contains(@id,"progressbar_disc")]')))
        wait()
        WebDriverWait(driver, 3600).until(EC.invisibility_of_element_located((By.XPATH, '//*[contains(@id,"pageloader")]')))
    else:
        WebDriverWait(driver, 3600).until(EC.invisibility_of_element_located((By.XPATH, '//*[contains(@id,"pageloader")]')))'''
pageloader()
WebDriverWait(driver,3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="fillallquote1"]/a')))
allquotes = driver.find_element(By.XPATH,'//*[@id="fillallquote1"]/a')
driver.execute_script("arguments[0].click();", allquotes)
#print("Element is visible? " + str(allquotes.is_displayed()))
#ActionChains(driver).move_to_element(allquotes).click(allquotes).perform()
pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="qterevall"]')))
allquotesearch = driver.execute_script('return document.querySelector("#qterevall > div.search_button input")')
driver.execute_script("arguments[0].click();", allquotesearch)
driver.execute_script('return document.querySelector("#qterevall > div.search_button input").value ="05984507"')
pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '(//button[@class="searched_button"])[2]')))
#time.sleep(20)
allsearch = driver.find_element(By.XPATH,'(//button[@class="searched_button"])[2]')
driver.execute_script("arguments[0].click();", allsearch)
time.sleep(8)
pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[5]/div[1]/div[6]/div[2]/div[1]/div/div/div[2]/div/div[1]/table/tbody/tr/td[2]/a')))
quoteclick = driver.find_element(By.XPATH,'/html/body/div[5]/div[1]/div[6]/div[2]/div[1]/div/div/div[2]/div/div[1]/table/tbody/tr/td[2]/a')
driver.execute_script("arguments[0].click();", quoteclick)
pageloader()
def refreshwait():
    time.sleep(5)
    i = 0
    while i==0:
        try:   
            if (str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none;' or str(driver.find_element(By.XPATH,'/html/body/div[2]').get_attribute("style")) == 'display: none; opacity: 0.2;'):
                i = 1
        except:
            exceptions.StaleElementReferenceException
            pass
refreshwait()
driver.refresh()

def addproductwait():
    pageloader()
    WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//button[@id="addproducts"]')))
addproductwait()
addproductstart = time.time()

addproduct = driver.find_element(By.XPATH,'//button[@id="addproducts"]')
driver.execute_script("arguments[0].click();", addproduct)

def bulkaddbtnwait():
    pageloader()
    WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//button[@id="bulkaddprd"]')))

salesorgid = driver.find_element(By.XPATH, '//*[@id="Primary_Panel"]/table/tbody/tr[2]/td[9]').text
print(salesorgid)

if salesorgid == '':    
    driver.refresh()
    addproductwait()
    addproductstart = time.time()
    addproduct = driver.find_element(By.XPATH,'//button[@id="addproducts"]')
    driver.execute_script("arguments[0].click();", addproduct)
    
def simpleproductwait():
    pageloader()
    WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="TH_ACTIONS"]/div[2]/button')))
simpleproductwait()
addproductend = time.time()
addproducttime = addproductend - addproductstart
print('add product time='+str(addproducttime))
sheet.cell(row=2,column=1).value = "Navigating to Product Catalog by Add Product"
sheet.cell(row=2,column=2).value = addproducttime
wait()
simpleproduct = driver.find_element(By.XPATH,'(//label[text()="Configurable"]/following::input)[2]')
driver.execute_script("arguments[0].click();", simpleproduct)
wait()
time.sleep(3)
simpleproductstart = time.time()
addsimpleproduct = driver.find_element(By.XPATH,'(//button[@id="BTN_add_quote"])[3]').click()
#print('add to cart')

bulkaddbtnwait()

pageloader()
simpleproductend = time.time()
addsimpleproducttime = simpleproductend - simpleproductstart
print('simple product time='+str(addsimpleproducttime))
sheet.cell(row=3,column=1).value = "Adding a Simple product"
sheet.cell(row=3,column=2).value = addsimpleproducttime

addconfigproduct = driver.find_element(By.XPATH,'//*[@id="addproducts"]').click()

pageloader()
WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//input[@id="PART"]')))
#addproduct.click()
prdsearch = driver.find_element(By.XPATH,'//input[@id="PART"]')
driver.execute_script("arguments[0].click();", prdsearch)
partno = "800TSI-SI"
wait()
configprdsearch = driver.find_element(By.XPATH,'//input[@id="PART"]').send_keys(partno)
wait()
partnosearch = driver.find_element(By.XPATH,'//*[@id="TH_ACTIONS"]/div[2]/button').click()
pageloader()
WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//td[@id="TD_PartNumber"]//a[1]')))


config_prdstart = time.time()
config_prd = driver.find_element(By.XPATH,'//td[@id="TD_PartNumber"]//a[1]').click()

pageloader()
WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//button[@id="get_pricing"]')))
config_prdend = time.time()
print('config_prd time='+str(config_prdend - config_prdstart))
wait()
getpricing = driver.find_element(By.XPATH,'//button[@id="get_pricing"]').click()


wait()
pageloader()
WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="cat_configure_save"]')))
saveproductstart = time.time()
saveproduct = driver.find_element(By.XPATH,'//*[@id="cat_configure_save"]').click()

pageloader()
WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="seginnerbnr"]/div[7]/div[2]')))
saveproductend = time.time()
saveconfigproducttime = saveproductend - saveproductstart
print('config_prd save time='+str(saveconfigproducttime))
sheet.cell(row=4,column=1).value = "Adding a Configurable product"
sheet.cell(row=4,column=2).value = saveconfigproducttime

netvalue = driver.find_element(By.XPATH, '//*[@id="seginnerbnr"]/div[7]/div[2]').text
print(netvalue)

bulkaddbtnwait()
bulkaddbtn = driver.find_element(By.XPATH,'//button[@id="bulkaddprd"]').click()

#pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="bulkaddsave"]')))
bulkpartnos = driver.find_element(By.XPATH,'//textarea[@id="partnumbers"]').click()
partnos = "1020508,TAKE3,GEN5,8040501,1225101,7170011,1450541,1213032, 7082249,48294,8040501"
wait()
insertpartnos = driver.find_element(By.XPATH,'//textarea[@id="partnumbers"]').send_keys(partnos)
wait()
bulkaddsavestart = time.time()
bulkaddsave = driver.find_element(By.XPATH,'//*[@id="bulkaddsave"]').click()

bulkaddbtnwait()
bulkaddsaveend = time.time()
bulkadd10products = bulkaddsaveend - bulkaddsavestart
print('bulk add time='+str(bulkadd10products))
sheet.cell(row=5,column=1).value = "Bulk Add products"
sheet.cell(row=5,column=2).value = bulkadd10products
wait()
discdropdown =  Select(driver.find_element(By.XPATH,'//div[@class="row prd_container"]//select[1]'))
wait()
discdropdown.select_by_visible_text('MON Discount View')
#mondiscview = driver.find_element(By.XPATH,'/html/body/div[9]/div[3]/div/div/div/div[4]/div[7]/select/option[4]').click()


pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//tr[1]/td/a[contains(@id,"YA9I")]')))
mondiscallcheck = driver.find_element(By.XPATH,'//input[@id="theadDiscountFirstChkbox"]').click()
wait()
YA9Idisc = driver.find_element(By.XPATH,'//tr[1]/td/a[contains(@id,"YA9I")]').click()
wait()
selectalldisc = driver.find_element(By.XPATH,'//span[text()="The record clicked"]/following::input').click()
wait()
discpercent = driver.find_element(By.XPATH,'(//span[text()="YA9I"])/following::div[2]/input')
discpercent.click()
wait()
#discpercentageclear = driver.find_element(By.XPATH,'/html/body/div[9]/div[21]/div/div/div[2]/div[1]/div[2]/div/input').clear()
discpercent.clear()
percentage = "10.0"
wait()
#discpercentage = driver.find_element(By.XPATH,'/html/body/div[9]/div[21]/div/div/div[2]/div[1]/div[2]/div/input')
discpercent.send_keys(percentage)
wait()
discpercentsavestart = time.time()
discpercentsave = driver.find_element(By.XPATH,'(//button[contains(@class,"btn btn-list-cust")]/following-sibling::button)[3]').click()
bulkaddbtnwait()
pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//abbr[@id="tot_quote_netValue"]')))
discpercentsaveend = time.time()
Applydiscount = discpercentsaveend - discpercentsavestart
print('applying discount time='+str(Applydiscount))
sheet.cell(row=6,column=1).value = "Applying Bulk Discount"
sheet.cell(row=6,column=2).value = Applydiscount
wait()
netvalue = driver.find_element(By.XPATH, '//abbr[@id="tot_quote_netValue"]').text
print(netvalue)

def approvalnodewait():
    pageloader()
    WebDriverWait(driver, 3600).until(EC.element_to_be_clickable((By.XPATH, '//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/approval_node.svg")]')))

if netvalue == '0.00':    
    reprice = driver.find_element(By.XPATH,'//*[@id="btn_Reprice"]').click()
    approvalnodewait()

approvalnodewait()
        
approvalnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/approval_node.svg")]').click()

pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="approvalIframe"]')))

approvaliframe = driver.find_element(By.XPATH, '//*[@id="approvalIframe"]')
wait()
driver.switch_to.frame(approvaliframe)
print("switched to approval iframe")


pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '(//textarea[@id="JUSTIFICATION_COMMENTS"])[2]')))
j_comment = driver.find_element(By.XPATH,'(//textarea[@id="JUSTIFICATION_COMMENTS"])[2]').click()
comment = "automation test"
wait()
justify = driver.find_element(By.XPATH,'(//textarea[@id="JUSTIFICATION_COMMENTS"])[2]').send_keys(comment)

pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//div[2]/button[1][@id="approval_act_btn"]')))
approvebuttonstart = time.time()
approvebutton = driver.find_element(By.XPATH,'//div[2]/button[1][@id="approval_act_btn"]')
ActionChains(driver).move_to_element(approvebutton).click(approvebutton).perform()


pageloader()
pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="approvalIframe"]')))
approvebuttonend = time.time()
ApprovingQuote = approvebuttonend - approvebuttonstart
print('approve time='+str(ApprovingQuote))
sheet.cell(row=7,column=1).value = "Self Approving Quote"
sheet.cell(row=7,column=2).value = ApprovingQuote
pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="approvalIframe"]')))
wait()
driver.switch_to.default_content()
wait()
print('switched to default content')


pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//img[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/order.svg")]')))
ordersnode = driver.find_element(By.XPATH,'//img[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/order.svg")]')
#ActionChains(driver).move_to_element(ordersnode).click().perform()
driver.execute_script("arguments[0].click();", ordersnode)
print('clicked orders node')

pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//button[text()="CREATE RFO"]')))
createRFO = driver.find_element(By.XPATH,'//button[text()="CREATE RFO"]').click()

def POnumberwait():
    pageloader()
    WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="po_number"]')))

POnumberwait()
ponum = "0987"
PO_number = driver.find_element(By.XPATH,'//input[@id="po_number"]').send_keys(ponum)
POnumberwait()
datepicker = driver.find_element(By.XPATH,'//input[@id="po_date"]').click()
wait()
datepicker = driver.find_element(By.XPATH,'//tr[3]/td[text()="10"]').click()
#print('date picked')
wait()
missingreason = "automation test"
PO_missing = driver.find_element(By.XPATH,'//input[@id="po_file_missing"]').send_keys(missingreason)
POnumberwait()
CreateRFOsubmitstart = time.time()
CreateRFOsubmit = driver.find_element(By.XPATH,'//button[@id="CreateRFO"]').click()

pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//p/a[@class="quote-title-span7"]')))
CreateRFOsubmitend = time.time()
CreatingRFO = CreateRFOsubmitend - CreateRFOsubmitstart
print('Create RFO time='+str(CreatingRFO))
sheet.cell(row=8,column=1).value = "Creating RFO"
sheet.cell(row=8,column=2).value = CreatingRFO
wait()
BacktoQuote = driver.find_element(By.XPATH,'//*[@id="Primary_Panel"]/table/tbody/tr[2]/td[7]/p/a')
driver.execute_script("arguments[0].click();", BacktoQuote)

pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/Revisions_node.svg")]')))
wait()
revisionsnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/Revisions_node.svg")]')
driver.execute_script("arguments[0].click();", revisionsnode)

def NewRevisionwait():
    pageloader()
    WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//button[@id="Newrevision"]')))
NewRevisionwait()
NewRevisionstart = time.time()
NewRevision = driver.find_element(By.XPATH,'//button[@id="Newrevision"]').click()
pageloader()
WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/product.svg")]')))
NewRevisionend = time.time()
Creating_New_Revision = NewRevisionend - NewRevisionstart
print('New Revision time='+str(Creating_New_Revision))
sheet.cell(row=9,column=1).value = "Creating New Revision"
sheet.cell(row=9,column=2).value = Creating_New_Revision


productsnode = driver.find_element(By.XPATH,'//*[contains(@src, "/mt/AGILENTTECHNOLOGIESINC_TST1/Additionalfiles/icons/product.svg")]').click()

#wait()
#driver.execute_script("window.scrollTo(0, 10)")
def deletechkboxwait():
    pageloader()
    WebDriverWait(driver, 3600).until(EC.visibility_of_element_located((By.XPATH, '//button[@id="btn_bulkDelete"]')))
deletechkboxwait()
wait()
prdpagination = driver.find_element(By.XPATH,'//div[@id="itemsPerPageDropdown"]').click()
wait()
prdpagination = driver.find_element(By.XPATH,'//a[contains(text(),"20")]').click()
deletechkboxwait()
deletechkbox = driver.find_element(By.XPATH,'//input[@class="custom inp_chkHdrMulDel"]').click()

time.sleep(5)
deletebutton = driver.find_element(By.XPATH,'//button[@id="btn_bulkDelete"]').click()

time.sleep(5)
deleteconfirmstart = time.time()
deleteconfirm = driver.find_element(By.XPATH,'//button[text()="DELETE"]').click()
#print('deleted')
addproductwait()
deleteconfirmend = time.time()
Bulkdelete12products = deleteconfirmend - deleteconfirmstart
print('bulk delete time='+str(Bulkdelete12products))
sheet.cell(row=10,column=1).value = "Bulk Deleting Products"
sheet.cell(row=10,column=2).value = Bulkdelete12products

sheet['A1'].font = Font(bold=True, size = 12)
sheet['A1'].alignment = Alignment(horizontal='center')
sheet['B1'].font = Font(bold=True, size = 12)
sheet['B1'].alignment = Alignment(horizontal='center')
'''for row in sheet['B']:
    cell = row[:10]
    cell.alignment = Alignment(horizontal='center')'''

workbook.save(excelpath)
time.sleep(10)
driver.quit()